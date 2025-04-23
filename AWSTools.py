import os
import configparser
import boto3
import logging
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import customtkinter as ctk
from tkinter import messagebox

__author__ = "Vikas Mahi   "

# Custom Theme Configuration
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# Constants
AWS_CONFIG_PATH = os.path.join(os.environ['USERPROFILE'], '.aws', 'config')
HARDCODED_REGIONS = [
    'us-east-1', 'us-east-2', 'us-west-1', 'us-west-2', 'af-south-1',
    'ap-east-1', 'ap-south-1', 'ap-northeast-1', 'ap-northeast-2',
    'ap-northeast-3', 'ap-southeast-1', 'ap-southeast-2', 'ap-southeast-3',
    'ca-central-1', 'eu-central-1', 'eu-west-1', 'eu-west-2', 'eu-west-3',
    'eu-south-1', 'eu-north-1', 'me-south-1', 'me-central-1', 'sa-east-1'
]

class ScrollableCheckboxFrame(ctk.CTkFrame):
    def __init__(self, master, title, **kwargs):
        super().__init__(master, **kwargs)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)  # Make checkbox frame expandable
        
        # Header with accent color
        self.header = ctk.CTkFrame(self, fg_color="#2B579A")
        self.header.grid(row=0, column=0, sticky="ew")
        
        self.title_label = ctk.CTkLabel(
            self.header,
            text=title,
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="white"
        )
        self.title_label.pack(padx=10, pady=5, anchor="w")
        
        self.checkbox_frame = ctk.CTkScrollableFrame(self)
        self.checkbox_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        
        self.checkboxes = {}
        self.all_var = ctk.BooleanVar()
        
        self.all_checkbox = ctk.CTkCheckBox(
            self.checkbox_frame, 
            text="Select All",
            variable=self.all_var,
            command=self.toggle_all,
            border_width=2,
            checkbox_width=20,
            checkbox_height=20
        )
        self.all_checkbox.pack(anchor="w", pady=(0, 8), padx=5)

    def toggle_all(self):
        state = self.all_var.get()
        for var in self.checkboxes.values():
            var.set(state)

    def add_items(self, items):
        for item in items:
            var = ctk.BooleanVar()
            checkbox = ctk.CTkCheckBox(
                self.checkbox_frame, 
                text=item, 
                variable=var,
                font=ctk.CTkFont(size=13),
                checkbox_width=18,
                checkbox_height=18,
                width=280  # Fix width to prevent text clipping
            )
            checkbox.pack(anchor="w", pady=2, padx=5)
            self.checkboxes[item] = var

    def get_selected(self):
        return [item for item, var in self.checkboxes.items() if var.get()]

class AWSReportTool(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("AWS Resource Reporter")
        self.geometry("1100x800")
        self.profiles = []
        self.account_name_map = {}

        # Configure window
        self.grid_rowconfigure(0, minsize=40)  # Fixed size for header
        self.grid_rowconfigure(1, weight=1)    # Main content expands
        self.grid_columnconfigure(0, weight=1)
        self.minsize(1000, 700)

        self.create_widgets()
        self.load_profiles()

    def create_widgets(self):
        # Header Section
        self.header_frame = ctk.CTkFrame(self, fg_color="#2B579A")
        self.header_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=0)
        
        # Use pack instead of grid for header content to avoid overlap issues
        # Main title
        self.app_title = ctk.CTkLabel(
            self.header_frame,
            text="AWS Resource Reporter",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="white"
        )
        self.app_title.pack(side="left", padx=20, pady=5)

        # Developer credit - more to the left
        self.dev_credit = ctk.CTkLabel(
            self.header_frame,
            text=f"Developed by: {__author__}",
            font=ctk.CTkFont(size=12, slant="italic"),
            text_color="#E0E0E0"
        )
        self.dev_credit.pack(side="right", padx=50, pady=5)  # Increased padding to move it left

        # Main Content
        self.main_frame = ctk.CTkFrame(self)
        self.main_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        self.main_frame.grid_rowconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)

        # Tab view
        self.tab_view = ctk.CTkTabview(self.main_frame)
        self.tab_view.grid(row=0, column=0, sticky="nsew")
        
        # Create tabs
        self.ec2_tab = self.tab_view.add("  🖥️ Stopped EC2  ")
        self.volume_tab = self.tab_view.add("  💾 Stale Volumes  ")
        
        # Status Bar
        self.status_bar = ctk.CTkFrame(self, height=30)
        self.status_bar.grid(row=2, column=0, sticky="ew")
        self.status_label = ctk.CTkLabel(
            self.status_bar,
            text="Ready",
            anchor="w",
            font=ctk.CTkFont(size=12)
        )
        self.status_label.pack(side="left", padx=20)

        # Configure Tabs
        self.setup_ec2_tab()
        self.setup_volume_tab()

    def setup_ec2_tab(self):
        # Configure grid with better proportions
        self.ec2_tab.grid_rowconfigure(0, weight=3)  # Main selection area
        self.ec2_tab.grid_rowconfigure(1, weight=1)  # Info/settings area
        self.ec2_tab.grid_columnconfigure(0, weight=1)
        self.ec2_tab.grid_columnconfigure(1, weight=1)

        # Profile selection
        self.ec2_account_frame = ScrollableCheckboxFrame(self.ec2_tab, "AWS Profiles")
        self.ec2_account_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # Region selection
        self.ec2_region_frame = ScrollableCheckboxFrame(self.ec2_tab, "AWS Regions")
        self.ec2_region_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        # Info panel that utilizes the space below
        self.ec2_info_frame = ctk.CTkFrame(self.ec2_tab)
        self.ec2_info_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
        
        # Add content to the info panel
        info_heading = ctk.CTkLabel(
            self.ec2_info_frame,
            text="EC2 Report Options",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        info_heading.grid(row=0, column=0, padx=15, pady=(10, 5), sticky="w")
        
        # Option 1: Include Tags
        self.include_tags_var = ctk.BooleanVar(value=True)
        include_tags_cb = ctk.CTkCheckBox(
            self.ec2_info_frame,
            text="Include instance tags in report",
            variable=self.include_tags_var,
            font=ctk.CTkFont(size=13)
        )
        include_tags_cb.grid(row=1, column=0, padx=20, pady=5, sticky="w")
        
        # Option 2: Group by Region
        self.group_by_region_var = ctk.BooleanVar(value=False)
        group_by_region_cb = ctk.CTkCheckBox(
            self.ec2_info_frame,
            text="Group results by region",
            variable=self.group_by_region_var,
            font=ctk.CTkFont(size=13)
        )
        group_by_region_cb.grid(row=2, column=0, padx=20, pady=5, sticky="w")
        
        # Days input (example of additional option)
        days_frame = ctk.CTkFrame(self.ec2_info_frame)
        days_frame.grid(row=1, column=1, padx=(50, 20), pady=5, sticky="w")
        
        days_label = ctk.CTkLabel(
            days_frame,
            text="Minimum days stopped:",
            font=ctk.CTkFont(size=13)
        )
        days_label.pack(side="left")
        
        self.ec2_days_entry = ctk.CTkEntry(
            days_frame,
            width=60,
            font=ctk.CTkFont(size=13),
            placeholder_text="0"
        )
        self.ec2_days_entry.pack(side="left", padx=10)

        # Generate button
        self.ec2_generate_btn = ctk.CTkButton(
            self.ec2_tab,
            text="🚀 Generate EC2 Report",
            command=self.run_ec2_report,
            font=ctk.CTkFont(size=14, weight="bold"),
            height=40
        )
        self.ec2_generate_btn.grid(row=2, column=0, columnspan=2, pady=20, padx=20)

    def setup_volume_tab(self):
        # Configure grid for volume tab
        self.volume_tab.grid_rowconfigure(0, weight=3)  # Main selection area
        self.volume_tab.grid_rowconfigure(1, weight=1)  # Info/settings area
        self.volume_tab.grid_columnconfigure(0, weight=1)
        self.volume_tab.grid_columnconfigure(1, weight=1)

        # Profile selection
        self.volume_account_frame = ScrollableCheckboxFrame(self.volume_tab, "AWS Profiles")
        self.volume_account_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # Region selection
        self.volume_region_frame = ScrollableCheckboxFrame(self.volume_tab, "AWS Regions")
        self.volume_region_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        # Info panel that utilizes the space below
        self.volume_info_frame = ctk.CTkFrame(self.volume_tab)
        self.volume_info_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
        
        # Add content to the info panel
        vol_info_heading = ctk.CTkLabel(
            self.volume_info_frame,
            text="Volume Report Options",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        vol_info_heading.grid(row=0, column=0, padx=15, pady=(10, 5), sticky="w")
        
        # Days input
        days_frame = ctk.CTkFrame(self.volume_info_frame)
        days_frame.grid(row=1, column=0, padx=20, pady=10, sticky="w")
        
        days_label = ctk.CTkLabel(
            days_frame,
            text="🗓️ Stale Volume Age (Days):",
            font=ctk.CTkFont(size=13)
        )
        days_label.pack(side="left")
        
        self.days_entry = ctk.CTkEntry(
            days_frame,
            width=80,
            font=ctk.CTkFont(size=13),
            placeholder_text="30"
        )
        self.days_entry.pack(side="left", padx=10)
        
        # Volume size option
        size_frame = ctk.CTkFrame(self.volume_info_frame)
        size_frame.grid(row=1, column=1, padx=(50, 20), pady=10, sticky="w")
        
        size_label = ctk.CTkLabel(
            size_frame,
            text="Minimum volume size (GB):",
            font=ctk.CTkFont(size=13)
        )
        size_label.pack(side="left")
        
        self.vol_size_entry = ctk.CTkEntry(
            size_frame,
            width=60,
            font=ctk.CTkFont(size=13),
            placeholder_text="0"
        )
        self.vol_size_entry.pack(side="left", padx=10)
        
        # Additional options
        self.include_snapshots_var = ctk.BooleanVar(value=False)
        include_snapshots_cb = ctk.CTkCheckBox(
            self.volume_info_frame,
            text="Include snapshot information",
            variable=self.include_snapshots_var,
            font=ctk.CTkFont(size=13)
        )
        include_snapshots_cb.grid(row=2, column=0, padx=20, pady=5, sticky="w")

        # Generate button
        self.volume_generate_btn = ctk.CTkButton(
            self.volume_tab,
            text="🚀 Generate Volume Report",
            command=self.run_volume_report,
            font=ctk.CTkFont(size=14, weight="bold"),
            height=40
        )
        self.volume_generate_btn.grid(row=2, column=0, columnspan=2, pady=20, padx=20)

    def load_profiles(self):
        config = configparser.ConfigParser()
        config.read(AWS_CONFIG_PATH)
        profiles = [section[8:] for section in config.sections() if section.startswith('profile ')]
        self.profiles = profiles
        self.account_name_map = {profile: profile for profile in profiles}
        
        # Populate both tabs
        self.ec2_account_frame.add_items(profiles)
        self.ec2_region_frame.add_items(HARDCODED_REGIONS)
        self.volume_account_frame.add_items(profiles)
        self.volume_region_frame.add_items(HARDCODED_REGIONS)

    def run_ec2_report(self):
        try:
            selected_profiles = self.ec2_account_frame.get_selected()
            selected_regions = self.ec2_region_frame.get_selected()
            
            if not selected_profiles or not selected_regions:
                messagebox.showwarning("Selection Required", "Select at least one profile and one region")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Stopped Instances"

            # Header styling
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(color="000000", bold=True)
            
            headers = ["Account ID", "Account Name", "Region", "Instance ID", "Name", "Stopped Since"]
            
            # Add tags column if selected
            if self.include_tags_var.get():
                headers.append("Tags")
                
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Data collection
            for profile in selected_profiles:
                session = boto3.Session(profile_name=profile)
                account_id = session.client('sts').get_caller_identity()['Account']
                account_name = self.account_name_map.get(profile, profile)

                for region in selected_regions:
                    try:
                        ec2 = session.client('ec2', region_name=region)
                        response = ec2.describe_instances(
                            Filters=[{'Name': 'instance-state-name', 'Values': ['stopped']}]
                        )

                        for reservation in response['Reservations']:
                            for instance in reservation['Instances']:
                                name = next(
                                    (tag['Value'] for tag in instance.get('Tags', []) 
                                    if tag['Key'] == 'Name'), ""
                                )
                                stopped_since = self.get_stopped_time(instance)
                                
                                # Prepare row data
                                row_data = [
                                    account_id, account_name, region,
                                    instance['InstanceId'], name, stopped_since
                                ]
                                
                                # Add tags if option selected
                                if self.include_tags_var.get():
                                    tags_str = "; ".join([f"{tag['Key']}={tag['Value']}" for tag in instance.get('Tags', [])])
                                    row_data.append(tags_str)
                                    
                                ws.append(row_data)
                    except Exception as e:
                        ws.append([account_id, account_name, region, f"Error: {str(e)}", "", ""])

            # Save file
            filename = self.save_report(wb, "ec2_stopped_instances")
            self.status_label.configure(text=f"EC2 Report saved: {filename}")
            messagebox.showinfo("Success", f"EC2 Report saved:\n{filename}")

        except Exception as e:
            self.status_label.configure(text=f"EC2 Error: {str(e)}")
            messagebox.showerror("Error", f"EC2 Report failed: {str(e)}")

    def run_volume_report(self):
        try:
            selected_profiles = self.volume_account_frame.get_selected()
            selected_regions = self.volume_region_frame.get_selected()
            days = int(self.days_entry.get() or "30")  # Default to 30 days if empty
            min_size = int(self.vol_size_entry.get() or "0")  # Default to 0 if empty
            cutoff_date = datetime.now() - timedelta(days=days)

            all_volumes = []

            for profile in selected_profiles:
                session = boto3.Session(profile_name=profile)
                account_id = session.client('sts').get_caller_identity()['Account']
                account_name = self.account_name_map.get(profile, profile)

                for region in selected_regions:
                    ec2 = session.client('ec2', region_name=region)
                    paginator = ec2.get_paginator('describe_volumes')
                    
                    for page in paginator.paginate():
                        for volume in page['Volumes']:
                            if volume['State'] == 'available' and volume['Size'] >= min_size:
                                create_time = volume['CreateTime'].replace(tzinfo=None)
                                if create_time < cutoff_date:
                                    instance_id = volume['Attachments'][0]['InstanceId'] if volume.get('Attachments') else 'N/A'
                                    
                                    volume_data = [
                                        account_id, account_name, region,
                                        volume['VolumeId'], volume['Size'],
                                        volume['VolumeType'], instance_id,
                                        create_time.strftime('%Y-%m-%d'),
                                        (datetime.now() - create_time).days
                                    ]
                                    
                                    # Add snapshot info if selected
                                    if self.include_snapshots_var.get():
                                        try:
                                            snapshots = ec2.describe_snapshots(
                                                Filters=[{'Name': 'volume-id', 'Values': [volume['VolumeId']]}]
                                            )
                                            snap_count = len(snapshots['Snapshots'])
                                            latest_snap = max(snapshots['Snapshots'], key=lambda x: x['StartTime']) if snap_count > 0 else None
                                            latest_date = latest_snap['StartTime'].strftime('%Y-%m-%d') if latest_snap else "No snapshots"
                                            
                                            volume_data.append(snap_count)
                                            volume_data.append(latest_date)
                                        except Exception:
                                            volume_data.append(0)
                                            volume_data.append("Error retrieving snapshots")
                                    
                                    all_volumes.append(volume_data)

            if all_volumes:
                filename = self.save_volume_report(all_volumes)
                self.status_label.configure(text=f"Volume Report saved: {filename}")
                messagebox.showinfo("Success", f"Volume Report saved:\n{filename}")
            else:
                self.status_label.configure(text="No stale volumes found")
                messagebox.showinfo("Info", "No stale volumes found")

        except Exception as e:
            self.status_label.configure(text=f"Volume Error: {str(e)}")
            messagebox.showerror("Error", f"Volume Report failed: {str(e)}")

    def get_stopped_time(self, instance):
        if 'StateTransitionReason' in instance:
            reason = instance['StateTransitionReason']
            if '(' in reason and ')' in reason:
                return reason.split('(')[-1].split(')')[0].strip()
        return "--"

    def save_report(self, wb, prefix):
        """Save report to Desktop/AWS_Reports folder"""
        desktop_path = os.path.join(os.environ['USERPROFILE'], 'Desktop')
        report_folder = os.path.join(desktop_path, "AWS_Reports")
        os.makedirs(report_folder, exist_ok=True)
        
        filename = os.path.join(
            report_folder,
            f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        wb.save(filename)
        return filename

    def save_volume_report(self, data):
        wb = Workbook()
        ws = wb.active
        ws.title = "Stale Volumes"

        headers = [
            'Account ID', 'Account Name', 'Region', 'Volume ID',
            'Size (GB)', 'Type', 'Last Instance', 'Created Date', 'Age (Days)'
        ]
        
        # Add snapshot headers if that option was selected
        if self.include_snapshots_var.get():
            headers.extend(['Snapshot Count', 'Latest Snapshot Date'])
        
        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill("solid", fgColor="D3D3D3")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row in data:
            ws.append(row)

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        return self.save_report(wb, "stale_volumes")

if __name__ == '__main__':
    app = AWSReportTool()
    app.mainloop()
